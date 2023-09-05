import openpyxl
from openpyxl import Workbook
import os

def writeExcelData(audio_file, text_file_name, ratingsInNumber):
    # Check if the Excel file exists, and create it if not
    if not os.path.exists('Reports/data.xlsx'):
        workbook = Workbook()
        sheet = workbook.active
        # Add headers to the columns
        sheet.append(["Audio Files", "Text Files", "Ratings"])
        workbook.save('Reports/data.xlsx')
    else:
        workbook = openpyxl.load_workbook('Reports/data.xlsx')

    # Rename the default sheet
    default_sheet = workbook.active
    default_sheet.title = 'Sales Rating'

    # Check if the sheet "Sales Rating" exists, and create it if not
    sheet_name = "Sales Rating"
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        # Add headers to the columns
        sheet.append(["Audio Files", "Text Files", "Ratings"])

    # Get the sheet
    sheet = workbook[sheet_name]

    # Add data to the Excel file
    sheet.append([audio_file, text_file_name, ratingsInNumber])

    # Save the changes
    workbook.save('Reports/data.xlsx')

